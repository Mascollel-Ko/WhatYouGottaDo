from __future__ import annotations

import csv
import hashlib
import json
from pathlib import Path
from typing import Iterable

import openpyxl


IDENTITY_SHEET = "05_IDENTITY_MASTER"
RUNTIME_SHEET = "07_RUNTIME_SNAPSHOT"
SCALAR_SHEET = "08_SCALAR_METADATA"
TRAINING_ROLE_SHEET = "18_TRAINING_ROLE"
PROGRAM_SLOT_SHEET = "19_PROGRAM_SLOT"
CONFLICT_SHEET = "21_CONFLICT_REPORT"
TIMING_SHEET = "28_PROGRAM_TIMING"
BOOTSTRAP_SHEET = "29_EXERCISE_BOOTSTRAP"

HISTORY_ONLY_STATUS = "HISTORY_ONLY_GENERIC"
PRODUCTION_ACTIVE = "PRODUCTION_ACTIVE"
HISTORY_COMPATIBILITY_ONLY = "HISTORY_COMPATIBILITY_ONLY"

DECISION_TOKENS = {"KEEP_CANONICAL", "PROPOSED_USER_APPROVED"}
HISTORY_ONLY_KEYS = {
    "single_leg_rdl",
    "ex_bd072cd",
}


def load_workbook(path: Path, *, read_only: bool = False):
    return openpyxl.load_workbook(path, read_only=read_only, data_only=False)


def sheet_rows(workbook, sheet_name: str) -> list[dict[str, str]]:
    sheet = workbook[sheet_name]
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return []
    headers = [text(value) for value in values[0]]
    return [
        {header: text(row[index] if index < len(row) else "") for index, header in enumerate(headers)}
        for row in values[1:]
        if any(text(value) for value in row)
    ]


def text(value) -> str:
    return "" if value is None else str(value).strip()


def index_by(rows: Iterable[dict[str, str]], key: str) -> dict[str, dict[str, str]]:
    result: dict[str, dict[str, str]] = {}
    for row in rows:
        value = row.get(key, "")
        if not value:
            raise ValueError(f"Blank {key}")
        if value in result:
            raise ValueError(f"Duplicate {key}: {value}")
        result[value] = row
    return result


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_csv(path: Path, headers: list[str], rows: Iterable[dict[str, str]]) -> int:
    materialized = list(rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as target:
        writer = csv.DictWriter(target, fieldnames=headers, lineterminator="\n", extrasaction="ignore")
        writer.writeheader()
        writer.writerows({header: text(row.get(header, "")) for header in headers} for row in materialized)
    return len(materialized)


def write_json(path: Path, value) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def parse_pipe(value: str) -> tuple[str, ...]:
    return tuple(token.strip() for token in value.split("|") if token.strip() and token.strip() != "NONE")
